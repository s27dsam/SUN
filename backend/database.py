"""
Handles MySQL connection and seeding the cancer mortality table from Excel.
"""

import os
import pymysql
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

# ── Connection config (reads from .env) ───────────────────────────────────────

DB_CONFIG = {
    "host":     os.getenv("DB_HOST", "localhost"),
    "port":     int(os.getenv("DB_PORT", 3306)),
    "database": os.getenv("DB_NAME", "health_dashboard"),
    "user":     os.getenv("DB_USER", "root"),
    "password": os.getenv("DB_PASSWORD", ""),
}

EXCEL_PATH = os.path.join(
    os.path.dirname(__file__),
    "aihw-can-122-CDiA-2023-Book-2a-Cancer-mortality-and-age-standardised-rates-by-age-5-year-groups.xlsx",
)
SHEET_NAME = "Table S2a.1"
DATA_START = 7   # first data row

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS cancer_mortality (
    id                  INT AUTO_INCREMENT PRIMARY KEY,
    data_type           TEXT,
    cancer_group_site   VARCHAR(255),
    year                INT,
    sex                 VARCHAR(20),
    age_group           VARCHAR(30),
    count               INT,
    age_specific_rate   FLOAT,
    asr_2001_aus_std    FLOAT,
    asr_2023_aus        FLOAT,
    asr_who             FLOAT,
    asr_segi            FLOAT,
    icd10_codes         VARCHAR(100)
)
"""

INSERT_SQL = """
INSERT INTO cancer_mortality
    (data_type, cancer_group_site, year, sex, age_group,
     count, age_specific_rate, asr_2001_aus_std,
     asr_2023_aus, asr_who, asr_segi, icd10_codes)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
"""

BATCH_SIZE = 5000

def get_connection():
    return pymysql.connect(**DB_CONFIG)

def _clean(value):
    """Return None for missing/placeholder values, otherwise the value itself."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in ("", ". .", ".."):
            return None
        return stripped
    return value


def _to_float(value):
    cleaned = _clean(value)
    if cleaned is None:
        return None
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _to_int(value):
    cleaned = _clean(value)
    if cleaned is None:
        return None
    try:
        return int(float(cleaned))
    except (ValueError, TypeError):
        return None


def get_all_skin_mortality_data():
    """
    Fetches year and death counts specifically for 'Melanoma of the skin'
    from the cancer_mortality table, including aggregated 'Persons' data.
    """
    conn = get_connection()
    with conn.cursor(pymysql.cursors.DictCursor) as cur:
        cur.execute("""
            SELECT year, SUM(count) AS deaths
            FROM cancer_mortality
            WHERE cancer_group_site = 'Melanoma of the skin' AND sex = 'Persons'
            GROUP BY year
            ORDER BY year
        """)
        result = cur.fetchall()
    conn.close()
    return result


def seed_database():
    """
    Drops the existing table, recreates it, and seeds it with data from the
    Excel file.
    """
    print("Connecting to database...")
    conn = get_connection()
    print("Connection successful.")

    with conn.cursor() as cur:
        print("Recreating 'cancer_mortality' table...")
        cur.execute("DROP TABLE IF EXISTS cancer_mortality")
        cur.execute(CREATE_TABLE_SQL)
        print("Table created successfully.")

        print(f"Loading workbook '{EXCEL_PATH}'...")
        workbook = load_workbook(filename=EXCEL_PATH, read_only=True)
        sheet = workbook[SHEET_NAME]
        print(f"Workbook loaded. Reading from sheet '{SHEET_NAME}'.")

        data_to_insert = []
        # Iterate over rows, starting from the first data row
        for row in sheet.iter_rows(min_row=DATA_START, values_only=True):
            # Skip rows where the first cell is empty, as they are separators
            if not _clean(row[0]):
                continue

            # Map cell values to their respective columns, with cleaning
            mapped_row = (
                _clean(row[0]),   # data_type
                _clean(row[1]),   # cancer_group_site
                _to_int(row[2]),  # year
                _clean(row[3]),   # sex
                _clean(row[4]),   # age_group
                _to_int(row[5]),  # count
                _to_float(row[6]),# age_specific_rate
                _to_float(row[7]),# asr_2001_aus_std
                _to_float(row[8]),# asr_2023_aus
                _to_float(row[9]),# asr_who
                _to_float(row[10]),# asr_segi
                _clean(row[11]),  # icd10_codes
            )
            data_to_insert.append(mapped_row)

            # Insert in batches to manage memory
            if len(data_to_insert) == BATCH_SIZE:
                cur.executemany(INSERT_SQL, data_to_insert)
                print(f"Inserted {len(data_to_insert)} rows...")
                data_to_insert = []

        # Insert any remaining rows
        if data_to_insert:
            cur.executemany(INSERT_SQL, data_to_insert)
            print(f"Inserted final {len(data_to_insert)} rows.")

    conn.commit()
    conn.close()
    print("Database seeding complete and connection closed.")


def main():
    """Main function to seed the database."""
    seed_database()

if __name__ == "__main__":
    main()
